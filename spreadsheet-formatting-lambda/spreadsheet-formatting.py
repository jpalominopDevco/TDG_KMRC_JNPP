import datetime
from openpyxl import load_workbook

spreadsheet_route = 'FOR-PS-03.xlsx'

workbook = load_workbook(spreadsheet_route)

worksheet = workbook['Detalle']

# worksheet['A1'].value = 'Timestamp'
# worksheet['B1'].value = 'Guia de seguimiento'
# worksheet['C1'].value = 'Devcognita evaluado'
# worksheet['D1'].value = 'Fecha del seguimiento'
# worksheet['E1'].value = 'Avance al plan de formacion'
# worksheet['F1'].value = 'Proactividad en el estudio'
# worksheet['G1'].value = 'Comunicacion en la sesion'
# worksheet['H1'].value = 'Desarrollo de acuerdos pendientes del seguimiento anterior'
# worksheet['I1'].value = 'Fit con el seniority'
# worksheet['J1'].value = 'Evolucion tecnica'
# worksheet['K1'].value = 'Capacidad recursiva y de investigacion'
# worksheet['L1'].value = 'Observaciones'
# worksheet['M1'].value = 'Promedio'

headers = [
    'Timestamp',
    'Guia de seguimiento',
    'Devcognita evaluado',
    'Fecha del seguimiento',
    'Avance al plan de formacion',
    'Proactividad en el estudio',
    'Comunicacion en la sesion',
    'Desarrollo de acuerdos pendientes del seguimiento anterior',
    'Fit con el seniority',
    'Evolucion tecnica',
    'Capacidad recursiva y de investigacion',
    'Observaciones',
    'Promedio'
]

for column, header in zip(worksheet.iter_cols(min_row=1, max_row=1, max_col=len(headers)), headers):
    column[0].value = header

for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=4):
    cell_c = row[0]
    if cell_c.value is not None:
        cell_c_coordinate = cell_c.coordinate
        if cell_c_coordinate != 'C1':
            worksheet[cell_c_coordinate] = cell_c.value.strip()

    cell_d = row[1]
    if cell_d.value is not None:
        cell_d_coordinate = cell_d.coordinate
        if cell_d_coordinate != 'D1':
            worksheet[cell_d_coordinate].number_format = 'mm/dd/yyyy'

workbook.save('FOR-PS-03.xlsx')
workbook.close()
