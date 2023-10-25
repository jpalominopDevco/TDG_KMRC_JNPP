# import datetime
from openpyxl import load_workbook
import logging
import boto3
import os

logger = logging.getLogger()
logger.setLevel(logging.INFO)

OUTPUT_FILE_NAME = "employees-formatted-data.xlsx"
S3_BUCKET_NAME = "staff-assessment-bucket"
S3_FILE_NAME = "raw_data/employees-raw-data.xlsx"
COLUMN_HEADERS = [
    'Timestamp',
    'Guia de seguimiento',
    'Devcognita evaluado',
    'Fecha del seguimiento',
    'Avance al plan de formacion',
    'Proactividad en el estudio',
    'Comunicacion en la sesion',
    'Desarrollo de acuerdos pendientes del seguimiento anterior',
    'Fit con el seniority',
    'Evolucion tecnica',
    'Capacidad recursiva y de investigacion',
    'Observaciones',
    'Promedio'
]
def run(event, context):
    # Obtiene el archivo de S3
    file_s3 = get_file_s3()

    # Ruta del archivo en la lambda
    output_file = '/tmp/' + OUTPUT_FILE_NAME

    # Guarda el archivo en el sistema de archivos local
    with open(output_file, 'wb') as file:
        file.write(file_s3)

    # Abre el archivo de Excel
    workbook = load_workbook(output_file)

    # Elimina la hoja Datos en el archivo de Excel
    if 'Datos' in workbook.sheetnames:
        workbook.remove(workbook['Datos'])

    # Selecciona la hoja Detalle como hoja de trabajo
    worksheet = workbook['Detalle']

    # Se modifican los headers en las columnas
    modify_column_headers(worksheet)

    # Se recorren las filas menos los headers
    current_row = 2
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=13):
        cell_c = row[2]
        cell_d = row[3]
        cell_m = row[12]

        if cell_c.value is not None:

            # Se recortan los espacios en blanco al inicio y al final del correo
            cell_c_coordinate = cell_c.coordinate
            if cell_c_coordinate != 'C1':
                worksheet[cell_c_coordinate] = cell_c.value.strip()

            # Se genera el promedio de cada empleado por sesión de seguimiento
            cell_m_coordinate = cell_m.coordinate
            if cell_m_coordinate != 'M1':
                worksheet[cell_m_coordinate] = f"=AVERAGE(E{current_row}:K{current_row})"
                current_row += 1
        else:
            break

        # Se agrega el formato especificado en la columna de fechas
        if cell_d.value is not None:
            cell_d_coordinate = cell_d.coordinate
            if cell_d_coordinate != 'D1':
                worksheet[cell_d_coordinate].number_format = 'mm/dd/yyyy'

    # Se salvan los cambios y cierra el archivo
    workbook.save(output_file)
    workbook.close()

    # Sube el archivo a S3
    upload_to_s3(output_file, S3_BUCKET_NAME, OUTPUT_FILE_NAME)

    # Elimina el archivo local
    os.remove(output_file)

def get_file_s3():
    
    s3 = boto3.client('s3')

    # Obtiene el archivo desde S3
    response = s3.get_object(Bucket=S3_BUCKET_NAME, Key=S3_FILE_NAME)

    # Lee el contenido del archivo
    file_content = response['Body'].read()

    logger.info('Archivo descargado')

    return file_content

def upload_to_s3(local_path, bucket_name, s3_key):
    s3_client = boto3.client('s3')
    s3_client.upload_file(local_path, bucket_name, 'formatted_data/' + s3_key)

def modify_column_headers(worksheet):
    # Se recorren las columnas y se modifica el header
    for column, header in zip(worksheet.iter_cols(min_row=1, max_row=1, max_col=len(COLUMN_HEADERS)), COLUMN_HEADERS):
        column[0].value = header